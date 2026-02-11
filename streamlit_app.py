import requests
import base64
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, numbers, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import uuid
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

# -------------------------------------------------
# HELPER FUNCTIONS (EDITOR LOGIC)
# -------------------------------------------------


def init_editor_structure(df):
    df = df.copy()

    if "_id" not in df.columns:
        df["_id"] = [str(uuid.uuid4())[:8] for _ in range(len(df))]

    if "_parentId" not in df.columns:
        df["_parentId"] = ""

    if "_order" not in df.columns:
        df["_order"] = list(range(len(df)))

    if "_group" not in df.columns:
        df["_group"] = (
            df.get("code", "").astype(str) + "|" +
            df.get("Power Type", "").astype(str)
        )

    return df


def get_sorted_editor_df(df):
    return df.sort_values(by=["_parentId", "_order"]).reset_index(drop=True)


def reorder_row(df, row_id, direction):
    df = df.copy()
    idx = df.index[df["_id"] == row_id]

    if len(idx) == 0:
        return df

    idx = idx[0]

    if direction == "up" and idx > 0:
        df.iloc[[idx, idx - 1]] = df.iloc[[idx - 1, idx]].values

    elif direction == "down" and idx < len(df) - 1:
        df.iloc[[idx, idx + 1]] = df.iloc[[idx + 1, idx]].values

    df["_order"] = list(range(len(df)))
    return df


def convert_type(df, row_id):
    df = df.copy()
    idx = df.index[df["_id"] == row_id]

    if len(idx) == 0:
        return df

    idx = idx[0]

    current_type = df.at[idx, "type"]

    if current_type == "item":
        df.at[idx, "type"] = "subitem"
    else:
        df.at[idx, "type"] = "item"

    return df


def delete_row(df, row_id):
    df = df.copy()
    df = df[df["_id"] != row_id]
    df = df.reset_index(drop=True)
    df["_order"] = list(range(len(df)))
    return df


def spread_row(df, source_id, target_ids):
    df = df.copy()

    source_row = df[df["_id"] == source_id]
    if source_row.empty:
        return df

    source_row = source_row.iloc[0]

    for target_id in target_ids:
        target_idx = df.index[df["_id"] == target_id]
        if len(target_idx) == 0:
            continue

        target_idx = target_idx[0]

        for col in df.columns:
            if col not in ["_id", "_parentId", "_order"]:
                df.at[target_idx, col] = source_row[col]

    return df


def editor_to_flat_df(df):
    df = df.copy()
    drop_cols = ["_id", "_parentId", "_order", "_group"]
    return df.drop(columns=[c for c in drop_cols if c in df.columns])


def sync_grid_edits(edited_df):
    return edited_df.copy()

st.set_page_config(layout="wide")

POWER_AUTOMATE_URL = st.secrets["power_automate"]["url"]

# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------
if "df" not in st.session_state:
    st.session_state.df = None
if "editor_df" not in st.session_state:
    st.session_state.editor_df = None
if "editor_initialized" not in st.session_state:
    st.session_state.editor_initialized = False
if "spread_mode" not in st.session_state:
    st.session_state.spread_mode = False
if "pending_delete" not in st.session_state:
    st.session_state.pending_delete = None
# -------------------------------------------------
# HEADER
# -------------------------------------------------
st.title("📊 Interactive Table Review & Price Analysis")


# -------------------------------------------------
# PDF UPLOAD → POWER AUTOMATE (HIGHEST PRIORITY)
# -------------------------------------------------
st.subheader("📄 Upload 3 Quote PDFs (Power Automate)")

pdfs = st.file_uploader(
    "Upload 1 or more PDF quotes",
    type=["pdf"],
    accept_multiple_files=True
)

if pdfs:  # at least one file uploaded
    if st.button("🚀 Process PDFs via Power Automate"):
        with st.spinner("Sending PDFs to Power Automate…"):

            files_payload = []
            for pdf in pdfs:
                encoded = base64.b64encode(pdf.read()).decode("ascii")
                files_payload.append({
                    "name": pdf.name,
                    "content": encoded
                })

            response = requests.post(
                POWER_AUTOMATE_URL,
                json={"files": files_payload},
                headers={
                    "Content-Type": "application/json"
                },
                timeout=600
            )

        if response.status_code != 200:
            st.error("Power Automate failed to process PDFs")
            st.stop()

        # Expecting base64 CSV back
        csv_bytes = base64.b64decode(response.json()["csv"])
        df = pd.read_csv(io.BytesIO(csv_bytes))

        # 🔑 HANDOFF POINT — everything else already works
        for col in ["type", "supplier", "brand", "code", "description", "Power Type"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        st.session_state.df = df
        st.session_state.editor_df = None
        st.session_state.editor_initialized = False

        st.session_state.current_job_path = None
        st.session_state.job_loaded_from_queue = False
        
        # 🔥 Store CSV bytes for download
        st.session_state.csv_bytes = csv_bytes

        st.success(f"✅ CSV generated from {len(pdfs)} PDF(s) and loaded")
        st.rerun()
else:
    st.info("Upload 1 or more PDFs to start processing")

# 🔥 AUTO-DOWNLOAD CSV BUTTON (appears after processing)
if "csv_bytes" in st.session_state and st.session_state.csv_bytes is not None:
    st.download_button(
        label="📥 Download Raw CSV",
        data=st.session_state.csv_bytes,
        file_name=f"quotes_raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv"
    )

# -------------------------------------------------
# UPLOAD FILE (MANUAL OVERRIDE)
# -------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload CSV or Excel (manual override)",
    type=["csv", "xlsx"]
)

if uploaded_file:
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    for col in ["type", "supplier", "brand", "code", "description", "Power Type"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # 🔹 Override queue state
    st.session_state.df = df.copy()
    st.session_state.editor_df = None
    st.session_state.editor_initialized = False
    st.session_state.spread_mode = False
    st.session_state.pending_delete = None
    st.session_state.current_job_path = None
    st.session_state.job_loaded_from_queue = False

    st.success("📤 Manual file loaded (queue overridden)")

# ==============================================================
# STEP 4: REPLACE your "✏️ Review Source Table" section with this
# ==============================================================
# Delete this block:
#   if st.session_state.df is not None:
#       st.subheader("✏️ Review Source Table")
#       st.session_state.df = st.data_editor(...)
#
# Paste the following:

if st.session_state.df is not None:

    # --- Initialize editor structure ---
    if not st.session_state.editor_initialized or st.session_state.editor_df is None:
        st.session_state.editor_df = init_editor_structure(st.session_state.df)
        st.session_state.editor_initialized = True

    edf = st.session_state.editor_df

    st.subheader("✏️ Quote Editor")

    # ──────────────────────────────────────────
    # CUSTOM CSS FOR AG GRID CONTAINER
    # ──────────────────────────────────────────
    st.markdown("""
    <style>
        /* AG Grid dark theme overrides */
        .ag-theme-streamlit {
            --ag-background-color: #0e1117;
            --ag-header-background-color: #1a1c24;
            --ag-odd-row-background-color: #0e1117;
            --ag-row-hover-color: rgba(77,171,247,0.06);
            --ag-border-color: #2d3140;
            --ag-header-foreground-color: #8b8fa3;
            --ag-foreground-color: #e8eaed;
            --ag-font-family: 'Segoe UI', sans-serif;
            --ag-font-size: 13px;
            --ag-row-height: 38px;
            --ag-header-height: 36px;
        }

        /* Subitem indent */
        .subitem-row { padding-left: 28px !important; }

        /* Item row bold */
        .item-row { font-weight: 600; }

        /* Winner column */
        .winner-cell {
            background-color: rgba(81,207,102,0.08) !important;
            border-left: 2px solid #51cf66 !important;
        }

        /* Type badges */
        .type-badge-item {
            background: rgba(77,171,247,0.12);
            color: #4dabf7;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
        }
        .type-badge-sub {
            background: rgba(139,143,163,0.12);
            color: #8b8fa3;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
        }

        /* Action buttons in grid */
        .grid-action-btn {
            background: transparent;
            border: 1px solid #2d3140;
            color: #8b8fa3;
            border-radius: 4px;
            width: 24px;
            height: 24px;
            cursor: pointer;
            font-size: 12px;
            margin: 0 1px;
        }
        .grid-action-btn:hover {
            background: #22252e;
            color: #e8eaed;
        }

        /* Compact buttons row */
        div[data-testid="stHorizontalBlock"] > div {
            padding: 0 2px;
        }
    </style>
    """, unsafe_allow_html=True)

    # ──────────────────────────────────────────
    # GROUP BY PRODUCT → RENDER AG GRID PER GROUP
    # ──────────────────────────────────────────
    product_groups = edf[edf["type"] == "item"][["_group"]].drop_duplicates()["_group"].tolist()

    for group_key in product_groups:
        code, power_type = group_key.split("|", 1)

        # Get all rows in this group (items + their subitems)
        group_item_ids = edf[
            (edf["type"] == "item") & (edf["_group"] == group_key)
        ]["_id"].tolist()

        group_rows = edf[
            (edf["_group"] == group_key) |
            (edf["_parentId"].isin(group_item_ids))
        ]

        # Sort properly: item then its children
        sorted_rows = get_sorted_editor_df(group_rows)

        supplier_count = len(sorted_rows[sorted_rows["type"] == "item"])

        # ── GROUP HEADER ──
        with st.expander(
            f"📦 **{code}** — {power_type} ({supplier_count} suppliers)",
            expanded=True
        ):
            # ── AG GRID CONFIGURATION ──
            gb = GridOptionsBuilder.from_dataframe(
                sorted_rows[["_id", "type", "supplier", "description", "price", "brand", "code", "Power Type", "_parentId", "_order"]]
            )

            # Column definitions
            gb.configure_column("_id", header_name="", hide=True)
            gb.configure_column("_parentId", hide=True)
            gb.configure_column("_order", hide=True)
            gb.configure_column("code", hide=True)
            gb.configure_column("Power Type", hide=True)
            gb.configure_column("brand", hide=True)

            # Type column with custom renderer
            type_renderer = JsCode("""
                function(params) {
                    if (params.value === 'item') {
                        return '<span style="background:rgba(77,171,247,0.12);color:#4dabf7;padding:2px 6px;border-radius:3px;font-size:10px;font-weight:600;text-transform:uppercase;">ITEM</span>';
                    } else {
                        return '<span style="background:rgba(139,143,163,0.12);color:#8b8fa3;padding:2px 6px;border-radius:3px;font-size:10px;font-weight:600;text-transform:uppercase;">SUB</span>';
                    }
                }
            """)
            gb.configure_column(
                "type",
                header_name="Type",
                width=70,
                cellRenderer=type_renderer,
                editable=False
            )

            # Supplier column
            gb.configure_column(
                "supplier",
                header_name="Supplier",
                width=140,
                editable=True
            )

            # Description with indent for subitems
            desc_renderer = JsCode("""
                function(params) {
                    var prefix = '';
                    if (params.data.type === 'subitem') {
                        prefix = '<span style="color:#5c6072;margin-right:4px;">↳</span>';
                    }
                    return prefix + params.value;
                }
            """)
            gb.configure_column(
                "description",
                header_name="Description",
                flex=2,
                editable=True,
                cellRenderer=desc_renderer
            )

            # Price column — editable with formatting
            price_formatter = JsCode("""
                function(params) {
                    if (params.value == null || params.value === '' || isNaN(params.value)) return '—';
                    return '$' + Number(params.value).toLocaleString('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2});
                }
            """)
            gb.configure_column(
                "price",
                header_name="Price",
                width=130,
                editable=True,
                type=["numericColumn"],
                valueFormatter=price_formatter
            )

            # Row styling: bold items, lighter subitems
            row_style = JsCode("""
                function(params) {
                    if (params.data.type === 'item') {
                        return {'font-weight': '600', 'background-color': 'rgba(77,171,247,0.03)'};
                    } else {
                        return {'color': '#8b8fa3', 'padding-left': '12px'};
                    }
                }
            """)

            # Selection for actions
            gb.configure_selection(
                selection_mode="single",
                use_checkbox=False
            )

            # Row dragging for reorder
            gb.configure_column(
                "type",
                rowDrag=True,
                rowDragText=JsCode("""
                    function(params) {
                        return params.rowNode.data.description;
                    }
                """)
            )

            # Grid options
            gb.configure_grid_options(
                rowDragManaged=True,
                animateRows=True,
                getRowStyle=row_style,
                suppressMoveWhenRowDragging=False,
                domLayout='autoHeight',
                rowHeight=38,
                headerHeight=36,
            )

            grid_options = gb.build()

            # ── RENDER AG GRID ──
            grid_response = AgGrid(
                sorted_rows[["_id", "type", "supplier", "description", "price", "brand", "code", "Power Type", "_parentId", "_order"]],
                gridOptions=grid_options,
                update_mode=GridUpdateMode.VALUE_CHANGED | GridUpdateMode.SELECTION_CHANGED,
                allow_unsafe_jscode=True,
                theme="streamlit",
                height=min(400, 36 + 38 * len(sorted_rows) + 10),
                key=f"grid_{group_key}",
                fit_columns_on_grid_load=True,
            )

            # ── SYNC EDITS BACK ──
            if grid_response and grid_response.data is not None:
                updated_data = grid_response.data
                for _, row in updated_data.iterrows():
                    rid = row.get("_id")
                    if rid and rid in edf["_id"].values:
                        for col in ["price", "description", "supplier"]:
                            if col in row.index:
                                edf.loc[edf["_id"] == rid, col] = row[col]
                st.session_state.editor_df = edf

            # ── SELECTED ROW ──
            selected = grid_response.selected_rows
            if selected is not None and len(selected) > 0:
                sel_row = selected.iloc[0] if hasattr(selected, 'iloc') else selected[0]
                sel_id = sel_row.get("_id", sel_row.get("_id", ""))
                sel_desc = sel_row.get("description", "")
                sel_type = sel_row.get("type", "")
            else:
                sel_id = None
                sel_desc = ""
                sel_type = ""

            # ── ACTION BUTTONS (compact row) ──
            st.markdown(f"<small style='color:#8b8fa3;'>Selected: <b>{sel_desc or 'Click a row'}</b></small>", unsafe_allow_html=True)

            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                if st.button("⬆ Up", key=f"up_{group_key}", disabled=sel_id is None, use_container_width=True):
                    st.session_state.editor_df = reorder_row(edf, sel_id, "up")
                    st.rerun()

            with col2:
                if st.button("⬇ Down", key=f"dn_{group_key}", disabled=sel_id is None, use_container_width=True):
                    st.session_state.editor_df = reorder_row(edf, sel_id, "down")
                    st.rerun()

            with col3:
                convert_label = "→ Sub" if sel_type == "item" else "→ Item"
                if st.button(f"🔄 {convert_label}", key=f"cv_{group_key}", disabled=sel_id is None, use_container_width=True):
                    st.session_state.editor_df = convert_type(edf, sel_id)
                    st.rerun()

            with col4:
                if st.button("📋 Spread", key=f"sp_{group_key}", disabled=sel_id is None, use_container_width=True):
                    st.session_state.spread_mode = sel_id
                    st.session_state._spread_group = group_key

            with col5:
                if st.button("🗑 Delete", key=f"dl_{group_key}", disabled=sel_id is None, use_container_width=True):
                    st.session_state.pending_delete = sel_id

            # ── DELETE CONFIRMATION ──
            if st.session_state.pending_delete and st.session_state.pending_delete == sel_id:
                del_row = edf[edf["_id"] == sel_id]
                if not del_row.empty:
                    del_row = del_row.iloc[0]
                    child_count = len(edf[edf["_parentId"] == sel_id])
                    warn_msg = f"Delete **{del_row['description']}**?"
                    if child_count > 0:
                        warn_msg += f" ⚠️ This will also delete **{child_count} subitems**."

                    st.warning(warn_msg)
                    c1, c2 = st.columns(2)
                    with c1:
                        if st.button("✅ Confirm Delete", key=f"cdel_{group_key}", type="primary", use_container_width=True):
                            st.session_state.editor_df = delete_row(edf, sel_id)
                            st.session_state.pending_delete = None
                            st.rerun()
                    with c2:
                        if st.button("Cancel", key=f"xdel_{group_key}", use_container_width=True):
                            st.session_state.pending_delete = None
                            st.rerun()

            # ── SPREAD PANEL ──
            if st.session_state.spread_mode and st.session_state.get("_spread_group") == group_key:
                st.info(f"📋 Spreading row to multiple items")
                all_items = edf[edf["type"] == "item"][["_id", "supplier", "code", "description"]].copy()
                all_items["label"] = all_items.apply(
                    lambda r: f"{r['supplier']} — {r['code']} — {r['description']}", axis=1
                )
                target_options = dict(zip(all_items["label"], all_items["_id"]))

                targets = st.multiselect(
                    "Copy as subitem to:",
                    options=list(target_options.keys()),
                    key=f"sptgt_{group_key}"
                )

                sc1, sc2 = st.columns(2)
                with sc1:
                    if st.button("✅ Spread Now", key=f"dosp_{group_key}", type="primary", use_container_width=True):
                        if targets:
                            target_ids = [target_options[t] for t in targets]
                            st.session_state.editor_df = spread_row(edf, st.session_state.spread_mode, target_ids)
                            st.session_state.spread_mode = False
                            st.success(f"Spread to {len(target_ids)} items!")
                            st.rerun()
                with sc2:
                    if st.button("Cancel", key=f"xsp_{group_key}", use_container_width=True):
                        st.session_state.spread_mode = False
                        st.rerun()

            # ── SUBTOTALS DISPLAY ──
            items_in_group = sorted_rows[sorted_rows["type"] == "item"]
            subtotals = []
            for _, itm in items_in_group.iterrows():
                item_price = float(itm["price"]) if pd.notna(itm["price"]) else 0
                children = sorted_rows[sorted_rows["_parentId"] == itm["_id"]]
                child_sum = children["price"].sum() if not children.empty else 0
                total = item_price + child_sum
                subtotals.append({"supplier": itm["supplier"], "total": total})

            if subtotals:
                st.markdown("**Supplier Totals (before tax):**")
                cols = st.columns(len(subtotals))
                min_total = min(s["total"] for s in subtotals)
                for i, s in enumerate(subtotals):
                    with cols[i]:
                        is_winner = s["total"] == min_total
                        label = f"🏆 {s['supplier']}" if is_winner else s["supplier"]
                        st.metric(label=label, value=f"${s['total']:,.2f}")

    # ──────────────────────────────────────────
    # ADD NEW ROW (global, below all groups)
    # ──────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### ➕ Add New Row")

    with st.expander("Add a new item or subitem", expanded=False):
        nc1, nc2 = st.columns(2)
        with nc1:
            new_type = st.radio("Type", ["item", "subitem"], horizontal=True, key="new_type")
        with nc2:
            new_supplier = st.text_input("Supplier", key="new_supplier")

        nc3, nc4, nc5 = st.columns(3)
        with nc3:
            new_brand = st.text_input("Brand", key="new_brand")
        with nc4:
            new_code = st.text_input("Code", key="new_code")
        with nc5:
            new_power = st.text_input("Power Type", key="new_power")

        nc6, nc7 = st.columns([3, 1])
        with nc6:
            new_desc = st.text_input("Description", key="new_desc")
        with nc7:
            new_price = st.number_input("Price", min_value=0.0, value=0.0, key="new_price")

        parent_id = ""
        if new_type == "subitem":
            all_items = edf[edf["type"] == "item"][["_id", "supplier", "code", "description"]].copy()
            all_items["label"] = all_items.apply(
                lambda r: f"{r['supplier']} — {r['code']} — {r['description']}", axis=1
            )
            parent_options = dict(zip(all_items["label"], all_items["_id"]))
            parent_label = st.selectbox("Parent Item", options=list(parent_options.keys()), key="new_parent")
            if parent_label:
                parent_id = parent_options[parent_label]

        if st.button("➕ Add Row", key="add_row_btn", type="primary"):
            new_id = str(uuid.uuid4())[:8]
            group_key = f"{new_code}|{new_power}"

            if new_type == "subitem" and parent_id:
                existing = edf[edf["_parentId"] == parent_id]
                new_order = int(existing["_order"].max() + 1) if not existing.empty else 0
            else:
                new_order = 0
                parent_id = ""

            new_row = {
                "_id": new_id,
                "_parentId": parent_id,
                "_order": new_order,
                "_group": group_key,
                "type": new_type,
                "supplier": new_supplier,
                "brand": new_brand,
                "code": new_code,
                "description": new_desc,
                "Power Type": new_power,
                "price": new_price,
            }
            st.session_state.editor_df = pd.concat(
                [edf, pd.DataFrame([new_row])], ignore_index=True
            )
            st.success(f"✅ Added: {new_desc}")
            st.rerun()

    # ──────────────────────────────────────────
    # SYNC BACK TO FLAT DF (feeds preview + Excel)
    # ──────────────────────────────────────────
    st.session_state.df = editor_to_flat_df(st.session_state.editor_df)

# -------------------------------------------------
# TAX INPUT
# -------------------------------------------------
st.subheader("💲 Tax Settings")
tax_percent = st.number_input("Tax Percentage", min_value=0.0, value=12.0)

# -------------------------------------------------
# HTML PREVIEW (EXCEL-STYLE)
# -------------------------------------------------
st.subheader("👀 Price Analysis Preview (HTML Table)")

def generate_html_table(df, tax_percent):
    tax_rate = tax_percent / 100

    html = """
    <div style="overflow-x:auto;">
    <style>
        table {
            border-collapse: collapse !important;
            width: 100%;
            margin-bottom: 40px;
            font-family: Arial, sans-serif;
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #bfbfbf;
        }

        th, td {
            border: 1px solid #bfbfbf !important;
            padding: 6px 8px;
            vertical-align: middle;
            text-align: left;
            background-clip: padding-box;
        }

        th {
            background-color: #dae9f8;
            font-weight: 600;
        }

        .total-row td {
            background-color: #fce4d6;
            font-weight: bold;
        }
    </style>
    """

    main_items = df[
        (df["type"] == "item") &
        df["Power Type"].notna() &
        (df["Power Type"] != "")
    ]

    for code, power_type in main_items[["code", "Power Type"]].drop_duplicates().values:

        items_for_code = df[
            (df["code"] == code) &
            (
                (df["Power Type"] == power_type) |
                (df["Power Type"].isna()) |
                (df["Power Type"] == "")
            ) &
            (df["type"].isin(["item", "subitem"]))
        ]

        suppliers = items_for_code["supplier"].unique()
        brand = items_for_code[items_for_code["type"] == "item"].iloc[0]["brand"]
        descriptions = items_for_code["description"].unique()

        body_rows = len(descriptions) + 2  # items + tax + total

        html += "<table>"

        # HEADER
        html += "<tr>"
        html += "<th>Details</th><th></th><th>QTY</th><th>Items</th>"
        for s in suppliers:
            html += f"<th>{s}</th>"
        html += "</tr>"

        totals = {s: 0 for s in suppliers}

        # FIRST ITEM ROW (with DETAILS)
        first_desc = descriptions[0]

        html += "<tr>"
        html += f"""
            <td rowspan="{body_rows}">
                <b>Brand</b><br>{brand}<br><br>
                <b>Code</b><br>{code}<br><br>
                <b>Power Type</b><br>{power_type}
            </td>
            <td rowspan="{body_rows}"></td>
            <td>1</td>
            <td>{first_desc}</td>
        """

        for s in suppliers:
            row = items_for_code[
                (items_for_code["supplier"] == s) &
                (items_for_code["description"] == first_desc)
            ]
            price = float(row["price"].iloc[0]) if not row.empty else 0
            totals[s] += price
            html += f"<td>${price:,.2f}</td>"

        html += "</tr>"

        # REMAINING ITEM ROWS
        for desc in descriptions[1:]:
            html += "<tr>"
            html += f"<td>1</td><td>{desc}</td>"

            for s in suppliers:
                row = items_for_code[
                    (items_for_code["supplier"] == s) &
                    (items_for_code["description"] == desc)
                ]
                price = float(row["price"].iloc[0]) if not row.empty else 0
                totals[s] += price
                html += f"<td>${price:,.2f}</td>"

            html += "</tr>"

        # TAX ROW
        html += "<tr>"
        html += "<td></td><td><b>Tax</b></td>"
        for _ in suppliers:
            html += f"<td>{tax_percent:.2f}%</td>"
        html += "</tr>"

        # TOTAL ROW
        html += "<tr class='total-row'>"
        html += "<td></td><td>Total</td>"
        for s in suppliers:
            total = totals[s] * (1 + tax_rate)
            html += f"<td>${total:,.2f}</td>"
        html += "</tr>"

        html += "</table>"

    html += "</div>"
    return html


# 🔥 RENDER HTML (LIVE, REACTIVE)
if (
    "df" in st.session_state
    and st.session_state.df is not None
    and not st.session_state.df.empty
):
    html = generate_html_table(st.session_state.df, tax_percent)
    st.markdown(html, unsafe_allow_html=True)
else:
    st.info("⬆️ Upload or generate data to see the price analysis preview.")

# -------------------------------------------------
# GENERATE FINAL EXCEL (MINIMALIST FORMATTING - RFQ STYLE)
# -------------------------------------------------
st.subheader("📥 Generate Final Excel")

if st.button("Generate Excel File"):
    df = st.session_state.df
    tax_rate = tax_percent / 100

    wb = Workbook()
    ws = wb.active
    ws.title = "Price Analysis"
    ws.sheet_view.showGridLines = False  # Clean minimalist look
    
    # --- MINIMALIST DESIGN TOKENS (MATCHING RFQ STYLE) ---
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Color palette matching RFQ Details
    HEADER_BLUE = PatternFill(start_color="288AD6", end_color="288AD6", fill_type="solid")
    DETAILS_BG = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    COLUMN_HEADER_BG = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    WINNER_BG = PatternFill(start_color="F2FAF2", end_color="F2FAF2", fill_type="solid")
    SPECS_BG = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    
    # Border styles
    SUBTLE_BORDER = Border(bottom=Side(style='thin', color="F0F0F0"))
    COLUMN_HEADER_BORDER = Border(bottom=Side(style='medium', color="E5E5E5"))
    TOTAL_BORDER = Border(top=Side(style='medium', color="288AD6"))
    
    # Text colors
    TEXT_PRIMARY = "1D1D1F"
    TEXT_SECONDARY = "86868B"
    WHITE = "FFFFFF"

    current_row = 1

    main_items = df[
        (df["type"] == "item") & 
        df["Power Type"].notna() & 
        (df["Power Type"] != "")
    ]

    for opt_idx, (code, power_type) in enumerate(
        main_items[["code", "Power Type"]].drop_duplicates().values, 1
    ):
        # Get all items for this code/power type combination
        items_for_code = df[
            (df["code"] == code) &
            (
                (df["Power Type"] == power_type) |
                (df["Power Type"].isna()) |
                (df["Power Type"] == "")
            ) &
            (df["type"].isin(["item", "subitem"]))
        ]

        suppliers = list(items_for_code["supplier"].unique())
        brand = items_for_code[items_for_code["type"] == "item"].iloc[0]["brand"]
        descriptions = list(items_for_code["description"].unique())

        # --- DETERMINE WINNER (LOWEST TOTAL PRICE) ---
        winner_supplier = ""
        min_total = float('inf')
        for sup in suppliers:
            sup_items = items_for_code[items_for_code["supplier"] == sup]
            total = sup_items["price"].sum() * (1 + tax_rate)
            if total < min_total:
                min_total = total
                winner_supplier = sup

        # === 1. OPTION TITLE (BLUE HEADER) ===
        title_row = current_row
        ws.row_dimensions[title_row].height = 40
        
        # Merge across all columns
        last_col = 6 + len(suppliers) - 1
        ws.merge_cells(
            start_row=title_row,
            start_column=2,
            end_row=title_row,
            end_column=last_col
        )
        
        title_cell = ws.cell(row=title_row, column=2, value=f"Option {opt_idx:02d}")
        title_cell.font = Font(name='Segoe UI', bold=True, size=14, color=WHITE)
        title_cell.fill = HEADER_BLUE
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

        # === 2. COLUMN HEADERS ===
        header_row = title_row + 1
        ws.row_dimensions[header_row].height = 28

        headers = ["DETAILS", "IMAGE", "QTY", "LINE ITEM"] + suppliers
        for i, h in enumerate(headers):
            col_idx = i + 2
            cell = ws.cell(row=header_row, column=col_idx, value=h.upper())
            cell.font = Font(name='Segoe UI', size=9, bold=True, color=TEXT_SECONDARY)
            cell.fill = WINNER_BG if h == winner_supplier else COLUMN_HEADER_BG
            cell.alignment = Alignment(
                horizontal="center" if col_idx >= 6 else "left",
                vertical="center"
            )
            cell.border = COLUMN_HEADER_BORDER

        # === 3. CONTENT ROWS ===
        start_data_row = header_row + 1
        num_item_rows = len(descriptions)
        num_body_rows = num_item_rows + 3  # +3 for total before tax, tax, and total rows

        # DETAILS column (merged vertically) - STOP BEFORE TOTAL ROW
        detail_val = f"BRAND\n{brand}\n\nCODE\n{code}\n\nPOWER\n{power_type}"
        d_cell = ws.cell(row=start_data_row, column=2, value=detail_val)
        d_cell.alignment = Alignment(
            wrap_text=True, 
            vertical="top", 
            horizontal="left",
            indent=1
        )
        d_cell.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        d_cell.fill = DETAILS_BG
        ws.merge_cells(
            start_row=start_data_row,
            start_column=2,
            end_row=start_data_row + num_item_rows + 1,  # Only through tax row
            end_column=2
        )
        
        # IMAGE placeholder (merged vertically) - STOP BEFORE TOTAL ROW
        img_cell = ws.cell(row=start_data_row, column=3, value="[ PHOTO ]")
        img_cell.alignment = Alignment(horizontal="center", vertical="center")
        img_cell.font = Font(name='Segoe UI', size=10, color="CCCCCC", italic=True)
        img_cell.fill = DETAILS_BG
        ws.merge_cells(
            start_row=start_data_row,
            start_column=3,
            end_row=start_data_row + num_item_rows + 1,  # Only through tax row
            end_column=3
        )

        # ITEM ROWS
        for idx, desc in enumerate(descriptions):
            r_num = start_data_row + idx
            ws.row_dimensions[r_num].height = 32

            # QTY column
            qty_cell = ws.cell(row=r_num, column=4, value=1)
            qty_cell.alignment = Alignment(horizontal="center", vertical="center")
            qty_cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY, bold=True)
            qty_cell.border = SUBTLE_BORDER

            # LINE ITEM (description)
            desc_cell = ws.cell(row=r_num, column=5, value=desc)
            desc_cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY)
            desc_cell.alignment = Alignment(vertical="center", horizontal="left")
            desc_cell.border = SUBTLE_BORDER

            # SUPPLIER PRICES (with formulas tied to QTY)
            qty_letter = get_column_letter(4)
            for s_idx, sup in enumerate(suppliers):
                col = 6 + s_idx

                # Get price for this supplier/description combo
                price_row = items_for_code[
                    (items_for_code["supplier"] == sup) &
                    (items_for_code["description"] == desc)
                ]
                price = float(price_row["price"].iloc[0]) if not price_row.empty else 0

                # Formula: =QTY * price
                cell = ws.cell(row=r_num, column=col, value=f"={qty_letter}{r_num}*{price}")
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = SUBTLE_BORDER
                cell.font = Font(name='Segoe UI', size=11, color=TEXT_PRIMARY)

                if sup == winner_supplier:
                    cell.fill = WINNER_BG

        # === 4. TOTAL BEFORE TAX ROW ===
        total_before_tax_row = start_data_row + num_item_rows
        ws.row_dimensions[total_before_tax_row].height = 32

        # Merge QTY and LINE ITEM columns for label
        ws.merge_cells(
            start_row=total_before_tax_row,
            start_column=4,
            end_row=total_before_tax_row,
            end_column=5
        )
        
        total_before_tax_label = ws.cell(row=total_before_tax_row, column=4, value="Total Before Tax")
        total_before_tax_label.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
        total_before_tax_label.alignment = Alignment(vertical="center", horizontal="left")
        total_before_tax_label.border = SUBTLE_BORDER

        for s_idx, sup in enumerate(suppliers):
            col = 6 + s_idx
            col_letter = get_column_letter(col)

            # Total Before Tax formula: SUM(items)
            tbt_cell = ws.cell(
                row=total_before_tax_row,
                column=col,
                value=f"=SUM({col_letter}{start_data_row}:{col_letter}{total_before_tax_row-1})"
            )
            tbt_cell.number_format = '$#,##0.00'
            tbt_cell.alignment = Alignment(horizontal="right", vertical="center")
            tbt_cell.font = Font(name='Segoe UI', size=11, bold=True, color=TEXT_PRIMARY)
            tbt_cell.border = SUBTLE_BORDER

            if sup == winner_supplier:
                tbt_cell.fill = WINNER_BG

        # === 5. TAX ROW ===
        tax_row = total_before_tax_row + 1
        ws.row_dimensions[tax_row].height = 28

        # Merge QTY and LINE ITEM columns for label
        ws.merge_cells(
            start_row=tax_row,
            start_column=4,
            end_row=tax_row,
            end_column=5
        )
        
        tax_label = ws.cell(row=tax_row, column=4, value=f"Tax ({int(tax_rate*100)}%)")
        tax_label.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        tax_label.alignment = Alignment(vertical="center", horizontal="left")
        tax_label.border = SUBTLE_BORDER

        for s_idx, sup in enumerate(suppliers):
            col = 6 + s_idx
            col_letter = get_column_letter(col)

            # Tax formula: Subtotal * tax_rate
            t_cell = ws.cell(
                row=tax_row,
                column=col,
                value=f"={col_letter}{total_before_tax_row}*{tax_rate}"
            )
            t_cell.number_format = '$#,##0.00'
            t_cell.alignment = Alignment(horizontal="right", vertical="center")
            t_cell.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
            t_cell.border = SUBTLE_BORDER

            if sup == winner_supplier:
                t_cell.fill = WINNER_BG

        # === 6. FINAL TOTAL ROW ===
        total_row = tax_row + 1
        ws.row_dimensions[total_row].height = 40
        
        # Merge DETAILS, IMAGE, QTY, and LINE ITEM columns (leave empty)
        ws.merge_cells(
            start_row=total_row,
            start_column=2,
            end_row=total_row,
            end_column=5
        )
        empty_cell = ws.cell(row=total_row, column=2, value="")
        empty_cell.fill = DETAILS_BG
        empty_cell.border = TOTAL_BORDER

        for s_idx, sup in enumerate(suppliers):
            col = 6 + s_idx
            col_letter = get_column_letter(col)

            # Total formula: Subtotal + Tax
            tot_cell = ws.cell(
                row=total_row,
                column=col,
                value=f"={col_letter}{total_before_tax_row}+{col_letter}{tax_row}"
            )
            tot_cell.font = Font(name='Segoe UI', bold=True, size=13, color=TEXT_PRIMARY)
            tot_cell.number_format = '$#,##0.00'
            tot_cell.alignment = Alignment(horizontal="right", vertical="center")
            tot_cell.border = TOTAL_BORDER

            if sup == winner_supplier:
                tot_cell.fill = WINNER_BG

        # === 7. SPECS & DESCRIPTION BLOCK ===
        specs_row = total_row + 1
        ws.row_dimensions[specs_row].height = 60

        # Merge across all columns
        ws.merge_cells(
            start_row=specs_row,
            start_column=2,
            end_row=specs_row,
            end_column=last_col
        )

        specs_content = ws.cell(
            row=specs_row,
            column=2,
            value="SPECS & DESCRIPTION\n\nEnter item specifications, dimensions, and technical details here..."
        )
        specs_content.font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
        specs_content.alignment = Alignment(
            wrap_text=True, 
            vertical="top",
            horizontal="left",
            indent=2
        )
        specs_content.fill = SPECS_BG
        specs_content.border = Border(bottom=Side(style='thin', color="F0F0F0"))

        # Move to next table (with spacing)
        current_row = specs_row + 3

    # === COLUMN WIDTH ADJUSTMENTS ===
    ws.column_dimensions['A'].width = 2   # Left margin
    ws.column_dimensions['B'].width = 18  # Details
    ws.column_dimensions['C'].width = 12  # Image
    ws.column_dimensions['D'].width = 8   # QTY
    ws.column_dimensions['E'].width = 35  # Line Item

    for i in range(len(suppliers)):
        ws.column_dimensions[get_column_letter(6+i)].width = 16

    # Save to downloadable file
    output = io.BytesIO()
    wb.save(output)

    st.download_button(
        "Download Excel",
        data=output.getvalue(),
        file_name=f"price_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )